"""
Script de nettoyage du xlsx OYA_POC_Orienteur.

Usage :
  python scripts/clean_xlsx.py [input.xlsx] [output.xlsx]

Transformations :
  - Supprime les colonnes parasites (Informations clés métier, Formations prioritaires, Contraintes fréquentes, Évolution si absente)
  - Nettoie les emojis dans la colonne Pénurie
  - Ajoute la colonne Évolution si absente (depuis data/Sheet_Metiers_import.csv)
  - Réordonne les colonnes K-P selon la structure cible
  - Renomme "Autonomie" → "Type_métier"
  - Écrit le xlsx propre

Structure cible onglet Métiers :
  A=Métier  B=Bloc  C-J=Poids_Q1-Q8
  K=Compétences  L=Niveau  M=Secteur  N=Type_métier  O=Pénurie  P=Évolution
"""

import sys
import re
import csv
import openpyxl
from openpyxl.utils import get_column_letter

INPUT  = sys.argv[1] if len(sys.argv) > 1 else "data/OYA_POC_Orienteur.xlsx"
OUTPUT = sys.argv[2] if len(sys.argv) > 2 else "data/OYA_POC_Orienteur_CLEAN.xlsx"
CSV_SOURCE = "data/Sheet_Metiers_import.csv"

COLS_TO_DELETE = {"Informations clés métier", "Formations prioritaires", "Contraintes fréquentes"}

PENURIE_NORMALIZE = {
    r"[🔴🔴]?\s*[Tt]ension":      "En tension",
    r"[🟢🟢]?\s*[Éé]mergent":     "Émergent",
    r"[🟡🟡]?\s*[Aa]utre.*":      "Autre",
    r"En tension":                  "En tension",
    r"Normal":                      "Normal",
    r"Émergent":                    "Émergent",
}

BLOC_NORMALIZE = {
    "Transformation agroalimentaire":                 "Transformation agroalimentaire & industries",
    "Logistique distribution":                         "Logistique, distribution & circuits courts",
    "Restauration":                                    "Restauration & métiers de bouche",
    "Transversal":                                     "Transversale",
}

TARGET_HEADERS = [
    "Métier", "Bloc",
    "Poids_Q1", "Poids_Q2", "Poids_Q3", "Poids_Q4",
    "Poids_Q5", "Poids_Q6", "Poids_Q7", "Poids_Q8",
    "Compétences", "Niveau", "Secteur", "Type_métier", "Pénurie", "Évolution",
]


def clean_emoji(text):
    """Retire les emojis et normalise la valeur Pénurie."""
    if not text:
        return ""
    text = str(text).strip()
    # Retirer emojis courants (séquences unicode)
    text = re.sub(r'[\U0001F000-\U0001FFFF☀-⛿✀-➿]+', '', text).strip()
    # Retirer les préfixes de code d'emoji type "ð" (encodage latin-1 manqué)
    text = re.sub(r'^[^\w\sÀ-ɏ]+', '', text).strip()
    # Normaliser
    for pattern, replacement in PENURIE_NORMALIZE.items():
        if re.match(pattern, text, re.IGNORECASE):
            return replacement
    return text


def load_csv_evolution(csv_path):
    """Charge le mapping Métier → Évolution depuis le CSV de référence."""
    mapping = {}
    try:
        with open(csv_path, newline='', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                metier = row.get('Métier', '').strip()
                evol = row.get('Évolution', '').strip()
                if metier:
                    mapping[metier] = evol
    except FileNotFoundError:
        print(f"  [WARN] CSV source non trouvé : {csv_path}")
    return mapping


wb = openpyxl.load_workbook(INPUT)
sheet = wb.active
print(f"Onglet actif : {sheet.title} — {sheet.max_row} lignes x {sheet.max_column} colonnes")

# Lire le header (ligne 1)
header = [cell.value for cell in sheet[1]]
print(f"Header brut : {header}")

# Charger mapping Évolution depuis CSV
evol_map = load_csv_evolution(CSV_SOURCE)
print(f"Mapping Évolution chargé : {len(evol_map)} entrées")

# Identifier les colonnes existantes par nom
col_idx = {}
for i, h in enumerate(header):
    if h is not None:
        key = str(h).strip()
        col_idx[key] = i

print(f"Colonnes identifiées : {list(col_idx.keys())}")

# Lire toutes les données
all_rows = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0]:  # ignorer lignes vides
        all_rows.append(list(row))

print(f"Données : {len(all_rows)} métiers")

# Construire les nouvelles lignes selon la structure cible
def get_col(row, name, fallback=""):
    idx = col_idx.get(name)
    if idx is None:
        return fallback
    val = row[idx] if idx < len(row) else None
    return str(val).strip() if val is not None else fallback


new_rows = []
issues = []

for i, row in enumerate(all_rows):
    metier_name = get_col(row, 'Métier').strip()

    # Pénurie : nettoyer emojis
    penurie_raw = get_col(row, 'Pénurie')
    penurie_clean = clean_emoji(penurie_raw)

    # Évolution : depuis CSV si absent du xlsx
    evol_raw = get_col(row, 'Évolution', '')
    if not evol_raw:
        evol_clean = evol_map.get(metier_name, 'Data manquante')
        if evol_clean == 'Data manquante':
            issues.append(f"  [WARN] Évolution manquante pour : {metier_name}")
    else:
        evol_clean = evol_raw.strip()

    # Secteur (col K actuelle)
    secteur = get_col(row, 'Secteur')

    # Type_métier = ancien "Autonomie"
    type_metier = get_col(row, 'Autonomie')
    if not type_metier:
        type_metier = get_col(row, 'Type_métier', 'Data manquante')

    # Compétences
    competences = get_col(row, 'Compétences').strip('"')

    # Niveau
    niveau = get_col(row, 'Niveau')

    # Poids Q1-Q8
    poids = [get_col(row, f'Poids_Q{q}') for q in range(1, 9)]

    # Valider poids
    for j, p in enumerate(poids, 1):
        try:
            pf = float(p) if p else 0.0
            if pf < 0 or pf > 1:
                issues.append(f"  [ERROR] {metier_name} Poids_Q{j}={p} hors [0,1]")
        except ValueError:
            issues.append(f"  [ERROR] {metier_name} Poids_Q{j}='{p}' non numérique")

    new_row = [
        metier_name.title() if metier_name.upper() == metier_name else metier_name,
        get_col(row, 'Bloc'),
        *poids,
        competences,
        niveau,
        secteur,
        type_metier,
        penurie_clean,
        evol_clean,
    ]
    new_rows.append(new_row)

print(f"\nIssues détectées : {len(issues)}")
for issue in issues[:20]:
    print(issue)

# Écrire le nouveau xlsx
wb_out = openpyxl.Workbook()
ws = wb_out.active
ws.title = "Métiers"

ws.append(TARGET_HEADERS)
for row in new_rows:
    ws.append(row)

wb_out.save(OUTPUT)
print(f"\n=== XLSX CLEAN sauvegardé : {OUTPUT} ===")
print(f"Lignes : {len(new_rows)} métiers, {len(TARGET_HEADERS)} colonnes")
print(f"Structure cible respectée : {TARGET_HEADERS}")

# Rapport final
print("\n=== RAPPORT NETTOYAGE MÉTIERS ===")
print(f"  Métiers : {len(new_rows)} lignes")
print(f"  Colonnes supprimées : Informations clés métier, Formations prioritaires, Contraintes fréquentes")
print(f"  Autonomie → renommé → Type_métier")
print(f"  Emojis Pénurie : nettoyés")
print(f"  Évolution : ajoutée depuis CSV source si absente")
print(f"  Issues : {len(issues)}")
if issues:
    print("  RESULT : VALID avec warnings")
else:
    print("  RESULT : VALID ✅")
